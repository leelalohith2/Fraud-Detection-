"""
Services Module
================
Higher-level business services that sit on top of the Hybrid AI Engine:

  * Fraud Forecasting      — hourly / daily / weekly fraud volume prediction
  * Alert System            — high-risk / fraud alerts (console, email, Telegram)
  * Report Engine           — PDF & Excel report generation
  * AI Fraud Copilot        — conversational investigation assistant
                               (uses OpenAI if OPENAI_API_KEY is set,
                               otherwise a deterministic template engine
                               so the whole platform works with zero keys)
"""
from __future__ import annotations

import io
import json
import smtplib
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from backend.config import get_logger, settings
from backend.database import insert_alert

logger = get_logger(__name__)


# ------------------------------------------------------------------
# Fraud Forecasting
# ------------------------------------------------------------------

def forecast_fraud_volume(scored_df: pd.DataFrame) -> dict:
    """Lightweight, dependency-free forecasting using weighted moving
    averages + day/hour seasonality extracted from historical data.
    Returns hourly (next 24h), daily (next 7d), and weekly (next 4w)
    predicted fraud counts."""
    df = scored_df.copy()
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df["is_flagged"] = (df["risk_category"].isin(["High Risk", "Fraud"])).astype(int)

    # --- Hourly seasonality profile (avg flagged txns per hour-of-day) ---
    hourly_profile = df.groupby(df["timestamp"].dt.hour)["is_flagged"].mean().reindex(range(24)).fillna(0)
    recent_hourly_volume = df[df["timestamp"] >= df["timestamp"].max() - pd.Timedelta(days=7)]
    base_hourly_count = max(1.0, len(recent_hourly_volume) / (7 * 24))

    now_hour = datetime.now().hour
    hourly_forecast = []
    for i in range(24):
        h = (now_hour + i) % 24
        predicted = base_hourly_count * (hourly_profile.get(h, hourly_profile.mean()) + 0.05)
        noise = np.random.normal(0, predicted * 0.08)
        hourly_forecast.append({
            "hour": f"{h:02d}:00",
            "predicted_fraud_count": max(0, round(predicted + noise, 1)),
        })

    # --- Daily trend (last 30 days -> simple linear trend + weekday seasonality) ---
    daily = df.groupby(df["timestamp"].dt.date)["is_flagged"].sum()
    if len(daily) >= 3:
        x = np.arange(len(daily))
        slope, intercept = np.polyfit(x, daily.values, 1)
    else:
        slope, intercept = 0.0, float(daily.mean() if len(daily) else 5)

    weekday_profile = df.groupby(df["timestamp"].dt.dayofweek)["is_flagged"].mean().reindex(range(7)).fillna(0)
    weekday_profile_norm = weekday_profile / (weekday_profile.mean() or 1)

    daily_forecast = []
    last_x = len(daily)
    for i in range(1, 8):
        d = datetime.now().date() + timedelta(days=i)
        trend_val = max(0, slope * (last_x + i) + intercept)
        seasonal_mult = weekday_profile_norm.get(d.weekday(), 1.0)
        predicted = trend_val * (0.6 + 0.4 * seasonal_mult)
        daily_forecast.append({
            "date": d.isoformat(),
            "predicted_fraud_count": max(0, round(predicted, 1)),
        })

    # --- Weekly forecast (aggregate of daily trend) ---
    weekly_forecast = []
    for w in range(1, 5):
        week_total = sum(
            max(0, slope * (last_x + (w - 1) * 7 + d) + intercept) for d in range(1, 8)
        )
        weekly_forecast.append({
            "week": f"Week +{w}",
            "predicted_fraud_count": max(0, round(week_total, 1)),
        })

    return {
        "hourly": hourly_forecast,
        "daily": daily_forecast,
        "weekly": weekly_forecast,
        "generated_at": datetime.now().isoformat(),
    }


# ------------------------------------------------------------------
# Alert System
# ------------------------------------------------------------------

def evaluate_alerts(scored_df: pd.DataFrame) -> pd.DataFrame:
    """Select transactions that cross alert thresholds."""
    alertable = scored_df[scored_df["risk_category"].isin(["High Risk", "Fraud"])].copy()
    return alertable


def send_email_alert(subject: str, body: str) -> bool:
    if not (settings.smtp_host and settings.alert_email_to):
        logger.info("Email alert skipped (SMTP not configured): %s", subject)
        return False
    try:
        msg = MIMEText(body)
        msg["Subject"] = subject
        msg["From"] = settings.smtp_user
        msg["To"] = settings.alert_email_to
        with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as server:
            server.starttls()
            server.login(settings.smtp_user, settings.smtp_password)
            server.send_message(msg)
        return True
    except Exception as e:
        logger.error("Email alert failed: %s", e)
        return False


def send_telegram_alert(message: str) -> bool:
    if not (settings.telegram_bot_token and settings.telegram_chat_id):
        logger.info("Telegram alert skipped (bot not configured)")
        return False
    try:
        url = f"https://api.telegram.org/bot{settings.telegram_bot_token}/sendMessage"
        resp = requests.post(url, json={"chat_id": settings.telegram_chat_id, "text": message}, timeout=5)
        return resp.status_code == 200
    except Exception as e:
        logger.error("Telegram alert failed: %s", e)
        return False


def raise_alerts(scored_df: pd.DataFrame, persist: bool = True) -> list[dict]:
    alertable = evaluate_alerts(scored_df)
    alerts = []
    for _, row in alertable.iterrows():
        severity = "CRITICAL" if row["risk_category"] == "Fraud" else "HIGH"
        message = (
            f"[{severity}] Transaction {row['transaction_id']} by {row['customer_id']} "
            f"flagged as {row['risk_category']} (score={row['final_risk_score']:.2f}, "
            f"amount={row.get('transaction_amount', 'N/A')})"
        )
        alerts.append({
            "transaction_id": row["transaction_id"], "customer_id": row["customer_id"],
            "severity": severity, "message": message,
        })
        if persist:
            try:
                insert_alert(row["transaction_id"], row["customer_id"], "risk_threshold", severity, message)
            except Exception as e:
                logger.warning("Could not persist alert: %s", e)
    return alerts


# ------------------------------------------------------------------
# Report Engine
# ------------------------------------------------------------------

def generate_pdf_report(scored_df: pd.DataFrame, title: str = "Fraud Intelligence Report") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], textColor=colors.HexColor("#0A0E14"))
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], textColor=colors.HexColor("#00D9FF"))

    elements = [
        Paragraph(title, title_style),
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]),
        Spacer(1, 0.6 * cm),
    ]

    total_txns = len(scored_df)
    fraud_txns = int((scored_df["risk_category"] == "Fraud").sum())
    high_risk = int((scored_df["risk_category"] == "High Risk").sum())
    money_protected = scored_df.loc[scored_df["risk_category"].isin(["Fraud", "High Risk"]), "transaction_amount"].sum()
    avg_risk = scored_df["final_risk_score"].mean()

    summary_data = [
        ["Metric", "Value"],
        ["Total Transactions", f"{total_txns:,}"],
        ["Fraud Detected", f"{fraud_txns:,}"],
        ["High Risk Flagged", f"{high_risk:,}"],
        ["Estimated Money Protected", f"₹{money_protected:,.0f}"],
        ["Average Risk Score", f"{avg_risk:.2%}"],
    ]
    summary_table = Table(summary_data, colWidths=[8 * cm, 8 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0E14")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements += [Paragraph("Executive Summary", heading_style), Spacer(1, 0.3 * cm), summary_table, Spacer(1, 0.6 * cm)]

    top_merchants = (
        scored_df[scored_df["risk_category"].isin(["Fraud", "High Risk"])]
        .groupby("merchant")["transaction_id"].count().sort_values(ascending=False).head(10)
    )
    merchant_data = [["Merchant", "Flagged Transactions"]] + [[m, str(c)] for m, c in top_merchants.items()]
    if len(merchant_data) > 1:
        merchant_table = Table(merchant_data, colWidths=[10 * cm, 6 * cm])
        merchant_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0E14")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements += [Paragraph("Top Risky Merchants", heading_style), Spacer(1, 0.3 * cm), merchant_table, Spacer(1, 0.6 * cm)]

    top_customers = (
        scored_df[scored_df["risk_category"].isin(["Fraud", "High Risk"])]
        .groupby("customer_id")["final_risk_score"].mean().sort_values(ascending=False).head(10)
    )
    cust_data = [["Customer ID", "Avg Risk Score"]] + [[c, f"{s:.2%}"] for c, s in top_customers.items()]
    if len(cust_data) > 1:
        cust_table = Table(cust_data, colWidths=[10 * cm, 6 * cm])
        cust_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0E14")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements += [Paragraph("High-Risk Customers", heading_style), Spacer(1, 0.3 * cm), cust_table]

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(scored_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill(start_color="0A0E14", end_color="0A0E14", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(["AI Fraud Intelligence Command Center — Report"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])

    total_txns = len(scored_df)
    fraud_txns = int((scored_df["risk_category"] == "Fraud").sum())
    high_risk = int((scored_df["risk_category"] == "High Risk").sum())
    money_protected = scored_df.loc[scored_df["risk_category"].isin(["Fraud", "High Risk"]), "transaction_amount"].sum()

    summary_rows = [
        ["Metric", "Value"],
        ["Total Transactions", total_txns],
        ["Fraud Detected", fraud_txns],
        ["High Risk Flagged", high_risk],
        ["Estimated Money Protected", round(money_protected, 2)],
        ["Average Risk Score", round(float(scored_df["final_risk_score"].mean()), 4)],
    ]
    for row in summary_rows:
        ws.append(row)
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font

    risk_dist = scored_df["risk_category"].value_counts()
    ws2 = wb.create_sheet("Risk Distribution")
    ws2.append(["Risk Category", "Count"])
    for cat, count in risk_dist.items():
        ws2.append([cat, int(count)])
    chart = BarChart()
    chart.title = "Risk Category Distribution"
    data = Reference(ws2, min_col=2, min_row=1, max_row=len(risk_dist) + 1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(risk_dist) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "D2")

    ws3 = wb.create_sheet("Transactions")
    export_cols = [
        "transaction_id", "customer_id", "transaction_amount", "timestamp", "merchant",
        "device_type", "location", "payment_method", "final_risk_score", "risk_category",
    ]
    export_cols = [c for c in export_cols if c in scored_df.columns]
    ws3.append(export_cols)
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for _, row in scored_df[export_cols].head(5000).iterrows():
        ws3.append(list(row))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ------------------------------------------------------------------
# AI Fraud Copilot
# ------------------------------------------------------------------

def _call_openai(prompt: str, context: str) -> Optional[str]:
    if not settings.openai_api_key:
        return None
    try:
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {settings.openai_api_key}"},
            json={
                "model": settings.openai_model,
                "messages": [
                    {"role": "system", "content": "You are a fraud investigation copilot for a bank SOC. "
                                                   "Be precise, concise, and reference the data context given."},
                    {"role": "user", "content": f"Context:\n{context}\n\nQuestion: {prompt}"},
                ],
                "max_tokens": 500,
            },
            timeout=15,
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"]
    except Exception as e:
        logger.warning("OpenAI copilot call failed, falling back to templates: %s", e)
    return None


def copilot_respond(question: str, scored_df: pd.DataFrame, engine=None) -> str:
    """Answer a fraud-investigation question. Tries OpenAI first (if
    configured), otherwise answers deterministically from the live
    scored dataframe so the copilot always works offline."""
    q = question.lower().strip()

    # Build a compact context summary for the LLM path
    context = json.dumps({
        "total_transactions": len(scored_df),
        "fraud_count": int((scored_df["risk_category"] == "Fraud").sum()),
        "high_risk_count": int((scored_df["risk_category"] == "High Risk").sum()),
        "top_cities": scored_df.groupby("location")["final_risk_score"].mean().sort_values(ascending=False).head(5).to_dict(),
    }, default=str)

    llm_answer = _call_openai(question, context)
    if llm_answer:
        return llm_answer

    # --- Deterministic fallback copilot ---
    if "why" in q and ("flagged" in q or "risk" in q or "fraud" in q):
        txn_id = None
        for token in question.split():
            if token.upper().startswith("TXN"):
                txn_id = token.upper()
        if txn_id and txn_id in scored_df["transaction_id"].values:
            row = scored_df[scored_df["transaction_id"] == txn_id].iloc[0]
            if engine is not None:
                exp = engine.explain_transaction(row)
                return (f"**{txn_id}** — Risk: {row['risk_category']} "
                        f"({row['final_risk_score']:.1%}).\n\n{exp['plain_english']}")
            return f"**{txn_id}** — Risk: {row['risk_category']} ({row['final_risk_score']:.1%})"
        return ("To explain a specific transaction, include its transaction ID (e.g. TXN...) in your question. "
                "In general, transactions are flagged when they combine several risk drivers: unusually high "
                "amounts, foreign or unfamiliar locations, rapid repeat transactions, device switching, or "
                "spending far outside a customer's normal pattern.")

    if "high-risk user" in q or "high risk user" in q or "risky customer" in q or "show high-risk" in q:
        top = (scored_df[scored_df["risk_category"].isin(["High Risk", "Fraud"])]
               .groupby("customer_id")["final_risk_score"].mean().sort_values(ascending=False).head(10))
        if top.empty:
            return "No high-risk customers found in the current dataset."
        lines = [f"- {cid}: {score:.1%} avg risk" for cid, score in top.items()]
        return "**Top high-risk customers:**\n" + "\n".join(lines)

    if "trend" in q:
        df = scored_df.copy()
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
        daily = df[df["risk_category"].isin(["High Risk", "Fraud"])].groupby(df["timestamp"].dt.date).size()
        if daily.empty:
            return "Not enough data yet to compute a fraud trend."
        direction = "rising" if daily.iloc[-1] >= daily.mean() else "stable/declining"
        return (f"Fraud activity over the last {len(daily)} days is **{direction}**. "
                f"Peak day: {daily.idxmax()} with {int(daily.max())} flagged transactions. "
                f"Average: {daily.mean():.1f} flagged transactions/day.")

    if "city" in q and ("highest" in q or "most" in q):
        city_scores = scored_df.groupby("location")["final_risk_score"].mean().sort_values(ascending=False)
        if city_scores.empty:
            return "No location data available."
        top_city = city_scores.index[0]
        return (f"**{top_city}** currently shows the highest average fraud risk "
                f"({city_scores.iloc[0]:.1%}), based on {scored_df[scored_df['location']==top_city].shape[0]} transactions.")

    if "predict" in q and ("tomorrow" in q or "next" in q):
        forecast = forecast_fraud_volume(scored_df)
        tomorrow = forecast["daily"][0] if forecast["daily"] else None
        if tomorrow:
            return (f"Forecast for {tomorrow['date']}: approximately "
                    f"**{tomorrow['predicted_fraud_count']:.0f} fraud/high-risk transactions** expected, "
                    f"based on recent trend and weekday seasonality.")
        return "Not enough historical data to generate a forecast yet."

    if "report" in q and "generate" in q:
        return ("I can generate a full PDF or Excel fraud report from the Reports Center — "
                "it will include executive KPIs, top risky merchants, and high-risk customers.")

    if "explain" in q and "risk score" in q:
        return (
            "The final risk score blends four independent signals:\n"
            "- **40% ML models** (XGBoost + LightGBM fraud probability)\n"
            "- **30% Anomaly detection** (Isolation Forest outlier score)\n"
            "- **20% Rule engine** (business rules like high amount, foreign txn, velocity)\n"
            "- **10% Behavioral analytics** (deviation from the customer's own normal pattern)\n\n"
            "Scores ≥90% are classified Fraud, ≥75% High Risk, ≥45% Medium Risk, below that Low Risk."
        )

    # Generic fallback
    fraud_count = int((scored_df["risk_category"] == "Fraud").sum())
    high_risk_count = int((scored_df["risk_category"] == "High Risk").sum())
    return (
        f"I can help investigate fraud patterns. Right now the system shows **{fraud_count} confirmed-fraud** "
        f"and **{high_risk_count} high-risk** transactions out of {len(scored_df):,} total. "
        "Try asking things like: 'Why was TXN... flagged?', 'Show high-risk users', "
        "'Show fraud trends', 'Which city has the highest fraud activity?', or 'Predict tomorrow's fraud volume.'"
    )
